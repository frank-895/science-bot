from pathlib import Path

import openpyxl
from science_bot.pipeline.resolution.planning import (
    ResolutionScratchpad,
    SearchAttempt,
    shortlist_candidate_files,
    tool_result_message,
    update_scratchpad_from_tool_result,
)
from science_bot.pipeline.resolution.tools.schemas import (
    CapsuleManifest,
    ColumnSearchResult,
    FileInfo,
)


def test_shortlist_candidate_files_prefers_family_keywords():
    manifest = CapsuleManifest(
        capsule_path="/tmp/capsule",
        total_size_bytes=3,
        files=[
            FileInfo(
                filename="clinical_table.csv",
                size_bytes=1,
                size_human="1 B",
                row_count=10,
                column_count=5,
                is_wide=False,
                file_type="csv",
            ),
            FileInfo(
                filename="random.csv",
                size_bytes=1,
                size_human="1 B",
                row_count=10,
                column_count=5,
                is_wide=False,
                file_type="csv",
            ),
        ],
    )

    candidates = shortlist_candidate_files(manifest, "aggregate")

    assert candidates[0].filename == "clinical_table.csv"


def test_shortlist_candidate_files_enriches_shortlisted_excel_candidates(
    tmp_path: Path,
):
    workbook = openpyxl.Workbook()
    workbook.active.title = "Tumor vs Normal"
    workbook.active.append(["protein", "gene", "log2FC", "adj.Pval"])
    workbook.active.append(["p1", "G1", 1.2, 0.01])
    workbook.create_sheet("Other")
    workbook.save(tmp_path / "Proteomic_data.xlsx")
    workbook.close()

    manifest = CapsuleManifest(
        capsule_path=str(tmp_path),
        total_size_bytes=1,
        files=[
            FileInfo(
                filename="Proteomic_data.xlsx",
                size_bytes=1,
                size_human="1 B",
                row_count=10,
                column_count=4,
                is_wide=False,
                file_type="excel",
            )
        ],
    )

    candidates = shortlist_candidate_files(
        manifest,
        "differential_expression",
        capsule_path=tmp_path,
    )

    assert candidates[0].sheet_names == ["Tumor vs Normal", "Other"]
    assert candidates[0].first_sheet_name == "Tumor vs Normal"
    assert candidates[0].first_sheet_columns == [
        "protein",
        "gene",
        "log2FC",
        "adj.Pval",
    ]


def test_shortlist_candidate_files_keeps_excel_candidate_when_enrichment_fails(
    tmp_path: Path,
):
    manifest = CapsuleManifest(
        capsule_path=str(tmp_path),
        total_size_bytes=1,
        files=[
            FileInfo(
                filename="Missing.xlsx",
                size_bytes=1,
                size_human="1 B",
                row_count=10,
                column_count=4,
                is_wide=False,
                file_type="excel",
            )
        ],
    )

    candidates = shortlist_candidate_files(
        manifest,
        "regression",
        capsule_path=tmp_path,
    )

    assert candidates[0].filename == "Missing.xlsx"
    assert candidates[0].sheet_names == []
    assert candidates[0].first_sheet_columns == []


def test_tool_result_message_reports_empty_search_results():
    message, truncated = tool_result_message("find_files_with_column", [])

    assert message == "No files with matching columns were found."
    assert truncated is False


def test_tool_result_message_reports_empty_filename_search_results():
    message, truncated = tool_result_message(
        "search_filenames",
        [],
        arguments={"query": "mafft"},
    )

    assert message == "No matching filenames were found for 'mafft'."
    assert truncated is False


def test_update_scratchpad_tracks_failed_search():
    scratchpad = ResolutionScratchpad(family="aggregate", question="question")

    update_scratchpad_from_tool_result(
        scratchpad=scratchpad,
        tool_name="find_files_with_column",
        arguments={"query": "gap"},
        result=[],
    )

    assert scratchpad.failed_searches == [
        SearchAttempt(
            tool_name="find_files_with_column",
            query="gap",
            outcome="no_matches",
        )
    ]
    assert scratchpad.last_tool_summary == "No files with matching columns were found."


def test_update_scratchpad_records_column_matches():
    scratchpad = ResolutionScratchpad(family="aggregate", question="question")
    result = ColumnSearchResult(
        filename="data.csv",
        query="gene",
        matches=["gene"],
        total_matches=1,
    )

    update_scratchpad_from_tool_result(
        scratchpad=scratchpad,
        tool_name="search_columns",
        arguments={"filename": "data.csv", "query": "gene"},
        result=result,
    )

    assert "data.csv" in scratchpad.selected_files
    assert scratchpad.column_evidence[0].columns == ["gene"]


def test_update_scratchpad_tracks_failed_filename_search():
    scratchpad = ResolutionScratchpad(family="aggregate", question="question")

    update_scratchpad_from_tool_result(
        scratchpad=scratchpad,
        tool_name="search_filenames",
        arguments={"query": "mafft"},
        result=[],
    )

    assert scratchpad.failed_searches == [
        SearchAttempt(
            tool_name="search_filenames",
            query="mafft",
            outcome="no_matches",
        )
    ]
    assert (
        scratchpad.last_tool_summary == "No matching filenames were found for 'mafft'."
    )


def test_update_scratchpad_tracks_zip_entries():
    scratchpad = ResolutionScratchpad(family="aggregate", question="question")

    from science_bot.pipeline.resolution.tools.schemas import ZipEntry, ZipManifest

    update_scratchpad_from_tool_result(
        scratchpad=scratchpad,
        tool_name="list_zip_contents",
        arguments={"zip_filename": "bundle.zip"},
        result=ZipManifest(
            zip_filename="bundle.zip",
            entries=[
                ZipEntry(
                    inner_path="run_eukaryota_odb10/full_table.tsv",
                    size_bytes=10,
                    file_type="tsv",
                    is_readable=True,
                )
            ],
        ),
    )

    assert scratchpad.known_zip_entries == {
        "bundle.zip": ["run_eukaryota_odb10/full_table.tsv"]
    }


def test_startup_excel_preview_is_not_added_to_known_columns(tmp_path: Path):
    workbook = openpyxl.Workbook()
    workbook.active.title = "Sheet1"
    workbook.active.append(["gene", "log2FC"])
    workbook.save(tmp_path / "results.xlsx")
    workbook.close()

    manifest = CapsuleManifest(
        capsule_path=str(tmp_path),
        total_size_bytes=1,
        files=[
            FileInfo(
                filename="results.xlsx",
                size_bytes=1,
                size_human="1 B",
                row_count=2,
                column_count=2,
                is_wide=False,
                file_type="excel",
            )
        ],
    )

    candidates = shortlist_candidate_files(
        manifest,
        "differential_expression",
        capsule_path=tmp_path,
    )
    scratchpad = ResolutionScratchpad(
        family="differential_expression",
        question="question",
        candidate_files=candidates,
    )

    assert scratchpad.known_columns == {}
    assert scratchpad.candidate_files[0].first_sheet_columns == ["gene", "log2FC"]
