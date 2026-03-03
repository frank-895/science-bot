from pathlib import Path

import openpyxl
import pytest
from science_bot.pipeline.resolution.tools.excel import list_excel_sheets


def test_list_excel_sheets_returns_workbook_sheet_names(tmp_path: Path):
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    workbook.create_sheet("Sheet2")
    file_path = tmp_path / "results.xlsx"
    workbook.save(file_path)
    workbook.close()

    sheet_names = list_excel_sheets(tmp_path, "results.xlsx")

    assert sheet_names == ["Summary", "Sheet2"]


def test_list_excel_sheets_uses_xlrd_for_xls(monkeypatch, tmp_path: Path):
    file_path = tmp_path / "results.xls"
    file_path.write_bytes(b"placeholder")

    called = {}

    class FakeWorkbook:
        def sheet_names(self):
            return ["Sheet1", "Sheet2"]

    def fake_open_workbook(path):
        called["path"] = path
        return FakeWorkbook()

    monkeypatch.setattr(
        "science_bot.pipeline.resolution.tools.excel.xlrd.open_workbook",
        fake_open_workbook,
    )

    assert list_excel_sheets(tmp_path, "results.xls") == ["Sheet1", "Sheet2"]
    assert called["path"] == str(file_path.resolve())


def test_list_excel_sheets_rejects_invalid_workbook(tmp_path: Path):
    file_path = tmp_path / "broken.xls"
    file_path.write_text("not really excel", encoding="utf-8")

    with pytest.raises(
        ValueError,
        match="does not appear to be a valid Excel workbook",
    ):
        list_excel_sheets(tmp_path, "broken.xls")
