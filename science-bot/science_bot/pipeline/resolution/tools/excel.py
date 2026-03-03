"""Excel-specific inspection tools for resolution."""

import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
import xlrd

from science_bot.pipeline.resolution.tools.reader import excel_engine, parse_filename


def list_excel_sheets(capsule_path: Path, filename: str) -> list[str]:
    """Return sheet names for an Excel file.

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Direct filename or zip-contained Excel path.

    Returns:
        list[str]: Sheet names in workbook order.
    """
    ref = parse_filename(capsule_path, filename)
    engine = excel_engine(ref.extension)
    try:
        if ref.zip_path is not None and ref.inner_path is not None:
            with zipfile.ZipFile(ref.zip_path) as archive:
                data = archive.read(ref.inner_path)
            if engine == "xlrd":
                workbook = xlrd.open_workbook(file_contents=data)
                return workbook.sheet_names()
            workbook = openpyxl.load_workbook(
                BytesIO(data),
                read_only=True,
                data_only=True,
            )
        else:
            if engine == "xlrd":
                workbook = xlrd.open_workbook(str(ref.file_path))
                return workbook.sheet_names()
            workbook = openpyxl.load_workbook(
                ref.file_path,
                read_only=True,
                data_only=True,
            )
    except Exception as exc:
        raise ValueError(
            f"File does not appear to be a valid Excel workbook: {ref.file_path}"
        ) from exc

    sheet_names = workbook.sheetnames
    workbook.close()
    return sheet_names
