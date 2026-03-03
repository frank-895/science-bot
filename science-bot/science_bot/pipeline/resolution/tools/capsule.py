"""Capsule inspection and data-loading tools for the resolution agent."""

import random
import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Literal, cast

import pandas as pd

from science_bot.pipeline.resolution.tools._reader import (
    _EXCEL_EXTENSIONS,
    _TABULAR_EXTENSIONS,
    count_rows,
    parse_filename,
    read_header,
    read_tabular,
)
from science_bot.pipeline.resolution.tools.schemas import (
    CapsuleManifest,
    ColumnInfo,
    ColumnSearchResult,
    ColumnStats,
    ColumnValues,
    ColumnValueSearchResult,
    FileInfo,
    FileSchema,
    RowSample,
    ZipEntry,
    ZipManifest,
)

WIDE_FILE_THRESHOLD: int = 200
MAX_SCHEMA_COLUMNS: int = 200
MAX_VALUE_COUNT: int = 50
READABLE_EXTENSIONS: frozenset[str] = frozenset(_TABULAR_EXTENSIONS)


def list_capsule_files(capsule_path: Path) -> CapsuleManifest:
    """Enumerate all readable files and zip containers in a capsule directory.

    Args:
        capsule_path: Absolute path to the capsule directory.

    Returns:
        CapsuleManifest: Sorted by file size descending.
    """
    infos: list[FileInfo] = []
    for p in sorted(capsule_path.iterdir()):
        if not p.is_file():
            continue
        ext = p.suffix.lower()
        if ext not in READABLE_EXTENSIONS and ext != ".zip":
            continue

        size_bytes = p.stat().st_size

        if ext == ".zip":
            infos.append(
                FileInfo(
                    filename=p.name,
                    size_bytes=size_bytes,
                    size_human=_human_size(size_bytes),
                    row_count=None,
                    column_count=None,
                    is_wide=None,
                    file_type="zip",
                )
            )
            continue

        file_type = _ext_to_file_type(ext)
        try:
            ref = parse_filename(capsule_path, p.name)
            columns = read_header(ref)
            col_count = len(columns)
            row_count = count_rows(ref)
        except Exception:
            col_count = None
            row_count = None

        infos.append(
            FileInfo(
                filename=p.name,
                size_bytes=size_bytes,
                size_human=_human_size(size_bytes),
                row_count=row_count,
                column_count=col_count,
                is_wide=col_count is not None and col_count > WIDE_FILE_THRESHOLD,
                file_type=file_type,
            )
        )

    infos.sort(key=lambda f: f.size_bytes, reverse=True)
    return CapsuleManifest(
        capsule_path=str(capsule_path),
        files=infos,
        total_size_bytes=sum(f.size_bytes for f in infos),
    )


def list_zip_contents(capsule_path: Path, zip_filename: str) -> ZipManifest:
    """List all entries inside a zip archive without extracting.

    Args:
        capsule_path: Absolute path to the capsule directory.
        zip_filename: Name of the zip file (e.g. ``"busco.zip"``).

    Returns:
        ZipManifest: All entries with size and readability metadata.
    """
    zip_path = (capsule_path / zip_filename).resolve()
    entries: list[ZipEntry] = []
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            ext = Path(info.filename).suffix.lower()
            file_type = _inner_ext_to_file_type(ext)
            entries.append(
                ZipEntry(
                    inner_path=info.filename,
                    size_bytes=info.file_size,
                    file_type=file_type,
                    is_readable=ext in READABLE_EXTENSIONS,
                )
            )
    return ZipManifest(zip_filename=zip_filename, entries=entries)


def list_excel_sheets(capsule_path: Path, filename: str) -> list[str]:
    """Return sheet names for an Excel file (direct or inside a zip).

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Direct filename or ``"archive.zip/file.xlsx"`` syntax.

    Returns:
        list[str]: Sheet names in workbook order.
    """
    import openpyxl

    ref = parse_filename(capsule_path, filename)
    if ref.zip_path is not None and ref.inner_path is not None:
        with zipfile.ZipFile(ref.zip_path) as zf:
            data = zf.read(ref.inner_path)
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(ref.file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def find_files_with_column(capsule_path: Path, query: str) -> list[ColumnSearchResult]:
    """Search column headers across all top-level readable files.

    Only top-level files are scanned (zip contents are not searched).

    Args:
        capsule_path: Absolute path to the capsule directory.
        query: Substring or regex pattern to match against column names.

    Returns:
        list[ColumnSearchResult]: One entry per file with at least one match.
    """
    results: list[ColumnSearchResult] = []
    for p in sorted(capsule_path.iterdir()):
        if not p.is_file():
            continue
        if p.suffix.lower() not in READABLE_EXTENSIONS:
            continue
        try:
            ref = parse_filename(capsule_path, p.name)
            matches = _filter_columns(read_header(ref), query)
        except Exception:
            continue
        if matches:
            results.append(
                ColumnSearchResult(
                    filename=p.name,
                    query=query,
                    matches=matches,
                    total_matches=len(matches),
                )
            )
    return results


def get_file_schema(capsule_path: Path, filename: str) -> FileSchema:
    """Return column-level schema derived from at most 5 data rows.

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Extended filename (supports zip and sheet syntax).

    Returns:
        FileSchema: Column names, dtypes, and sample values.
    """
    ref = parse_filename(capsule_path, filename)
    df = cast(pd.DataFrame, read_tabular(ref, nrows=5))
    all_columns = list(df.columns)
    col_count = len(all_columns)

    truncated = col_count > MAX_SCHEMA_COLUMNS
    columns_to_describe = all_columns[:MAX_SCHEMA_COLUMNS]
    df = df[columns_to_describe]

    row_count = count_rows(ref)

    column_infos: list[ColumnInfo] = []
    for col in columns_to_describe:
        series = df[col]
        null_count = int(series.isna().sum())
        sample_vals = [_sanitize_value(v) for v in series.tolist()]
        dtype_label = _infer_dtype_label(series)
        column_infos.append(
            ColumnInfo(
                name=col,
                dtype=dtype_label,
                sample_values=sample_vals,
                null_count_in_sample=null_count,
            )
        )

    return FileSchema(
        filename=filename,
        row_count=row_count,
        column_count=col_count,
        columns=column_infos,
        columns_truncated=truncated,
        max_schema_columns=MAX_SCHEMA_COLUMNS,
    )


def search_columns(capsule_path: Path, filename: str, query: str) -> ColumnSearchResult:
    """Search column headers in a single file (header-only, no data loaded).

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Extended filename.
        query: Substring or regex pattern.

    Returns:
        ColumnSearchResult: Matching column names.
    """
    ref = parse_filename(capsule_path, filename)
    headers = read_header(ref)
    matches = _filter_columns(headers, query)
    return ColumnSearchResult(
        filename=filename,
        query=query,
        matches=matches,
        total_matches=len(matches),
    )


def get_column_values(
    capsule_path: Path,
    filename: str,
    column: str,
    max_values: int = MAX_VALUE_COUNT,
) -> ColumnValues:
    """Return unique values observed in a single column.

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Extended filename.
        column: Column name to inspect.
        max_values: Maximum distinct values to return.

    Returns:
        ColumnValues: Unique values with truncation flag.
    """
    ref = parse_filename(capsule_path, filename)
    df = cast(pd.DataFrame, read_tabular(ref, usecols=[column]))
    series = df[column]
    unique_vals = series.dropna().unique().tolist()
    unique_count = len(unique_vals)
    truncated = unique_count > max_values
    values = [_sanitize_value(v) for v in unique_vals[:max_values]]
    dtype_label = _infer_dtype_label(series)
    return ColumnValues(
        filename=filename,
        column=column,
        dtype=dtype_label,
        unique_count=unique_count,
        values=values,
        truncated=truncated,
    )


def get_column_stats(capsule_path: Path, filename: str, column: str) -> ColumnStats:
    """Return descriptive statistics for a single column.

    Numeric columns get min/max/mean/std. String columns get top-10 most common.

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Extended filename.
        column: Column name to inspect.

    Returns:
        ColumnStats: Aggregated statistics.
    """
    ref = parse_filename(capsule_path, filename)
    df = cast(pd.DataFrame, read_tabular(ref, usecols=[column]))
    series = df[column]
    dtype_label = _infer_dtype_label(series)
    row_count = len(series)
    null_count = int(series.isna().sum())
    unique_count = int(series.nunique(dropna=True))

    col_min = col_max = col_mean = col_std = None
    most_common = None

    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().sum() > 0 and dtype_label in {"integer", "float"}:
        col_min = float(numeric.min())
        col_max = float(numeric.max())
        col_mean = float(numeric.mean())
        col_std = float(numeric.std(ddof=1)) if len(numeric.dropna()) > 1 else None
    else:
        counts = series.dropna().astype(str).value_counts().head(10)
        most_common = [(str(k), int(v)) for k, v in counts.items()]

    return ColumnStats(
        filename=filename,
        column=column,
        dtype=dtype_label,
        row_count=row_count,
        null_count=null_count,
        unique_count=unique_count,
        min=col_min,
        max=col_max,
        mean=col_mean,
        std=col_std,
        most_common=most_common,
    )


def search_column_for_value(
    capsule_path: Path,
    filename: str,
    column: str,
    query: str,
    max_matches: int = 50,
) -> ColumnValueSearchResult:
    """Find distinct values in a column that contain the query string.

    Case-insensitive substring match on string-cast values.

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Extended filename.
        column: Column name to search.
        query: Substring to search for (case-insensitive).
        max_matches: Maximum distinct matching values to return.

    Returns:
        ColumnValueSearchResult: Matching values with truncation flag.
    """
    ref = parse_filename(capsule_path, filename)
    df = cast(pd.DataFrame, read_tabular(ref, usecols=[column]))
    series = df[column]
    query_lower = query.lower()

    unique_vals = series.dropna().unique()
    matches: list[str | int | float | bool] = []
    for val in unique_vals:
        if query_lower in str(val).lower():
            matches.append(_sanitize_value(val))  # type: ignore[arg-type]

    total_matches = len(matches)
    truncated = total_matches > max_matches
    return ColumnValueSearchResult(
        filename=filename,
        column=column,
        query=query,
        matches=matches[:max_matches],
        total_matches=total_matches,
        truncated=truncated,
    )


def get_row_sample(
    capsule_path: Path,
    filename: str,
    columns: list[str],
    n: int = 10,
    random_sample: bool = False,
) -> RowSample:
    """Return a head or random sample of named columns.

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Extended filename.
        columns: Column names to include. Must be non-empty.
        n: Number of rows to return.
        random_sample: If True, use reservoir sampling for a random sample.

    Returns:
        RowSample: Sampled rows with metadata.

    Raises:
        ValueError: If columns is empty.
    """
    if not columns:
        raise ValueError("columns must be non-empty.")

    ref = parse_filename(capsule_path, filename)
    total_rows = count_rows(ref)

    if not random_sample:
        df = cast(pd.DataFrame, read_tabular(ref, usecols=columns, nrows=n))
        rows = _df_to_row_dicts(df)
        return RowSample(
            filename=filename,
            columns=columns,
            rows=rows,
            total_rows_in_file=total_rows,
            sample_size=len(rows),
            sampled_randomly=False,
        )

    # Reservoir sampling (Knuth Algorithm R) over streamed chunks.
    reservoir: list[dict] = []
    row_index = 0

    chunked = read_tabular(ref, usecols=columns, chunksize=1000)
    if isinstance(chunked, pd.DataFrame):
        chunks = [chunked]
    else:
        chunks = chunked

    for chunk in chunks:
        for _, row in chunk.iterrows():
            row_dict = {col: _sanitize_value(row[col]) for col in columns}
            if row_index < n:
                reservoir.append(row_dict)
            else:
                j = random.randint(0, row_index)
                if j < n:
                    reservoir[j] = row_dict
            row_index += 1

    return RowSample(
        filename=filename,
        columns=columns,
        rows=reservoir,
        total_rows_in_file=total_rows,
        sample_size=len(reservoir),
        sampled_randomly=True,
    )


def load_dataframe(
    capsule_path: Path,
    filename: str,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    """Load a tabular file into a DataFrame for downstream execution.

    This is the only tool that returns a DataFrame; it is intended for
    ExecutionPayload assembly and not for LLM inspection.

    Args:
        capsule_path: Absolute path to the capsule directory.
        filename: Extended filename.
        columns: Columns to load. If None, all columns are loaded — but
            raises ValueError for wide files (>WIDE_FILE_THRESHOLD columns).

    Returns:
        pd.DataFrame: Loaded data.

    Raises:
        ValueError: If columns is None and the file is wide, or if any
            requested column does not exist.
    """
    ref = parse_filename(capsule_path, filename)

    if columns is None:
        headers = read_header(ref)
        if len(headers) > WIDE_FILE_THRESHOLD:
            raise ValueError(
                f"File '{filename}' has {len(headers)} columns "
                f"(>{WIDE_FILE_THRESHOLD}). Specify 'columns' explicitly."
            )
        return cast(pd.DataFrame, read_tabular(ref))

    # Validate requested columns exist
    headers = read_header(ref)
    missing = [c for c in columns if c not in headers]
    if missing:
        raise ValueError(f"Columns not found in '{filename}': {missing}")

    return cast(pd.DataFrame, read_tabular(ref, usecols=columns))


def _human_size(size_bytes: int) -> str:
    """Format a byte count as a human-readable string."""
    for unit in ("B", "KB", "MB", "GB"):
        if size_bytes < 1024:
            return f"{size_bytes} {unit}"
        size_bytes //= 1024
    return f"{size_bytes} TB"


def _ext_to_file_type(ext: str) -> Literal["csv", "tsv", "excel"]:
    if ext in _EXCEL_EXTENSIONS:
        return "excel"
    if ext == ".tsv" or ext == ".tab":
        return "tsv"
    return "csv"


def _inner_ext_to_file_type(
    ext: str,
) -> Literal["csv", "tsv", "excel", "json", "other"]:
    if ext in _EXCEL_EXTENSIONS:
        return "excel"
    if ext in {".tsv", ".tab"}:
        return "tsv"
    if ext == ".csv":
        return "csv"
    if ext == ".json":
        return "json"
    return "other"


def _filter_columns(columns: list[str], query: str) -> list[str]:
    """Return columns matching query as substring first, then as regex."""
    lower_q = query.lower()
    substring_matches = [c for c in columns if lower_q in c.lower()]
    if substring_matches:
        return substring_matches
    try:
        pattern = re.compile(query, re.IGNORECASE)
        return [c for c in columns if pattern.search(c)]
    except re.error:
        return []


def _infer_dtype_label(
    series: pd.Series,
) -> Literal["string", "integer", "float", "boolean", "mixed"]:
    """Infer a human-readable dtype label from a pandas Series."""
    dtype = series.dtype
    if pd.api.types.is_bool_dtype(dtype):
        return "boolean"
    if pd.api.types.is_integer_dtype(dtype):
        return "integer"
    if pd.api.types.is_float_dtype(dtype):
        return "float"
    # object dtype — probe a sample
    sample = series.dropna().head(20)
    if sample.empty:
        return "string"
    types = {type(v) for v in sample}
    if types <= {int}:
        return "integer"
    if types <= {float}:
        return "float"
    if types <= {bool}:
        return "boolean"
    if types <= {int, float}:
        return "float"
    if len(types) > 1:
        return "mixed"
    return "string"


def _sanitize_value(
    val: object,
) -> str | int | float | bool | None:
    """Coerce a value to a JSON-serialisable scalar."""
    import math

    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return val
    return str(val)


def _df_to_row_dicts(
    df: pd.DataFrame,
) -> list[dict[str, str | int | float | bool | None]]:
    return [
        {col: _sanitize_value(row[col]) for col in df.columns}
        for _, row in df.iterrows()
    ]
